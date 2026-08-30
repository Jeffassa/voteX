"""Import en masse d'étudiants depuis un fichier Excel ESATIC.

Format attendu (réel ESATIC) :
- Le fichier .xlsx contient PLUSIEURS feuilles
- Le nom de chaque feuille = libellé de la classe (ex: "MP2I A", "SRIT 1B", "L3 GL")
- Un en-tête institutionnel précède souvent les vraies colonnes — on cherche
  automatiquement la première ligne contenant "matricule" comme en-têtes.
- Colonnes attendues (souples sur la casse / accents / pluriel) :
    - matricule (OBLIGATOIRE, format XX-ESATICNNNNAA)
    - nom / noms       (OBLIGATOIRE)
    - prenom / prenoms (OBLIGATOIRE)
    - genre / sexe     (optionnel — M/F/H/Masculin/Féminin/Homme/Femme/X)

Les étudiants importés ont password_hash=NULL → ils devront s'inscrire via
/register pour activer leur compte (matricule + nom + mot de passe).
"""

import io
import logging
from dataclasses import dataclass
from typing import Any

from fastapi import BackgroundTasks
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.matricule import is_valid_matricule, normalize_name
from app.models import ClassRoom, Student
from app.models.student import Gender, UserRole
from app.schemas.student_import import ImportReport, ImportRowResult
import secrets
from app.services import email_service


logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class PendingActivationEmail:
    """Un mail d'activation à envoyer une fois l'import commité."""

    to_email: str
    voter_name: str
    activation_code: str


COLUMN_ALIASES = {
    "matricule": {"matricule", "matricules", "mat", "matric", "id", "n°matricule", "numatricule"},
    "first_name": {"prenom", "prenoms", "firstname", "first_name", "first"},
    "last_name": {"nom", "noms", "lastname", "last_name", "last", "name", "nomdefamille"},
    "gender": {"genre", "gender", "sexe", "sex"},
    "email": {"email", "e-mail", "mail", "courriel"},
}

# Mapping inverse (alias normalisé → clé canonique) pré-calculé pour rapidité
_ALIAS_TO_CANONICAL: dict[str, str] = {
    normalize_name(alias): canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}

GENDER_NORMALIZATION = {
    "m": Gender.MALE, "masculin": Gender.MALE, "homme": Gender.MALE, "h": Gender.MALE, "male": Gender.MALE,
    "f": Gender.FEMALE, "feminin": Gender.FEMALE, "femme": Gender.FEMALE, "female": Gender.FEMALE,
    "x": Gender.OTHER, "autre": Gender.OTHER, "nb": Gender.OTHER, "other": Gender.OTHER,
}

# Nombre max de lignes à scanner pour trouver la vraie ligne d'en-têtes
HEADER_SEARCH_DEPTH = 20


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    return normalize_name(str(s))


def _row_header_score(row: tuple) -> tuple[int, dict[str, int]]:
    """Compte combien de cellules de cette ligne ressemblent à des en-têtes connus.
    Retourne (score, mapping nom canonique → index)."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        norm = _norm_header(cell)
        if not norm:
            continue
        canonical = _ALIAS_TO_CANONICAL.get(norm)
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return len(mapping), mapping


def _detect_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """Scanne les premières lignes pour trouver celle qui ressemble le plus à
    une ligne d'en-têtes (contient `matricule` + au moins un autre champ).
    Retourne (index de la ligne, mapping) ou None si rien trouvé."""
    best: tuple[int, int, dict[str, int]] | None = None  # (score, idx, mapping)
    for idx, row in enumerate(rows[:HEADER_SEARCH_DEPTH]):
        score, mapping = _row_header_score(row)
        # On exige au minimum la colonne matricule + 1 autre (nom OU prenom)
        if "matricule" in mapping and score >= 2:
            if best is None or score > best[0]:
                best = (score, idx, mapping)
    if best is None:
        return None
    return best[1], best[2]


def _resolve_class_by_sheet_name(db: Session, sheet_name: str) -> ClassRoom | None:
    """Le nom de feuille est le libellé de la classe.
    Tentatives :
      1. Match sur le nom COMPLET normalisé (insensible accents/casse/espaces)
         vs `{level} {name}` ou juste `name` en base.
      2. Si la feuille commence par un préfixe niveau (L1/L2/L3/M1/M2), split.
    """
    cleaned = sheet_name.strip()
    if not cleaned:
        return None
    norm_sheet = normalize_name(cleaned)

    all_classes = db.query(ClassRoom).all()

    # Match sur le nom complet ou juste le nom (le plus permissif d'abord)
    for c in all_classes:
        if normalize_name(f"{c.level} {c.name}") == norm_sheet:
            return c
    for c in all_classes:
        if normalize_name(c.name) == norm_sheet:
            return c

    # Tentative split level + name (ex: "L3 Génie Logiciel")
    parts = cleaned.split(maxsplit=1)
    if len(parts) >= 2:
        level, name_hint = parts[0].upper(), parts[1].strip()
        for c in all_classes:
            if c.level.upper() == level and normalize_name(c.name) == normalize_name(name_hint):
                return c
        # Permissif : nom contient le hint
        for c in all_classes:
            if c.level.upper() == level and normalize_name(name_hint) in normalize_name(c.name):
                return c

    return None


def _auto_create_class(db: Session, *, sheet_name: str, default_level: str) -> ClassRoom:
    """Crée à la volée une classe avec le nom de feuille comme `name` et le
    niveau passé en paramètre. Le champ `field` est rempli avec le même nom."""
    cleaned_name = sheet_name.strip()
    classroom = ClassRoom(
        name=cleaned_name,
        level=default_level.upper(),
        field=cleaned_name,
    )
    db.add(classroom)
    db.flush()
    return classroom


def _parse_gender(value: str | None) -> Gender | None:
    """Normalise avant lookup : les fichiers ESATIC écrivent "Féminin",
    "MASCULIN", "Homme "… et un simple .lower() laissait passer les accents,
    ce qui perdait silencieusement le genre de toutes les étudiantes."""
    if not value:
        return None
    return GENDER_NORMALIZATION.get(normalize_name(str(value)))


def _process_sheet(
    db: Session,
    *,
    sheet_name: str,
    rows: list[tuple],
    seen_in_file: set[str],
    dry_run: bool,
    auto_create_classes: bool = False,
    default_level: str | None = None,
    pending_emails: list["PendingActivationEmail"] | None = None,
) -> tuple[list[ImportRowResult], int, int]:
    """Traite une feuille — retourne (results, created, skipped)."""
    if pending_emails is None:
        pending_emails = []
    results: list[ImportRowResult] = []
    created = 0
    skipped = 0

    if not rows:
        return results, 0, 0

    detected = _detect_header_row(rows)
    if detected is None:
        results.append(
            ImportRowResult(
                row=0, status="error",
                message=(
                    f"Feuille '{sheet_name}' : aucune ligne d'en-tête trouvée. "
                    f"Vérifie qu'il y a bien une ligne avec 'matricule' + 'nom' + 'prénom' "
                    f"dans les {HEADER_SEARCH_DEPTH} premières lignes."
                ),
            )
        )
        return results, 0, 0

    header_row_idx, header_map = detected
    data_start = header_row_idx + 1  # 0-indexed offset dans rows

    classroom = _resolve_class_by_sheet_name(db, sheet_name)
    if not classroom:
        if auto_create_classes and default_level:
            classroom = _auto_create_class(
                db, sheet_name=sheet_name, default_level=default_level
            )
            results.append(
                ImportRowResult(
                    row=0, status="ok",
                    message=(
                        f"[{sheet_name}] Classe créée automatiquement : "
                        f"{classroom.level} {classroom.name}"
                    ),
                )
            )
        else:
            results.append(
                ImportRowResult(
                    row=0, status="error",
                    message=(
                        f"Feuille '{sheet_name}' : classe introuvable. "
                        f"Coche 'Créer auto les classes manquantes' (avec un niveau) "
                        f"ou crée-la manuellement dans /admin/classes."
                    ),
                )
            )
            return results, 0, 0

    if "first_name" not in header_map or "last_name" not in header_map:
        missing = []
        if "first_name" not in header_map:
            missing.append("prénom")
        if "last_name" not in header_map:
            missing.append("nom")
        results.append(
            ImportRowResult(
                row=header_row_idx + 1, status="error",
                message=(
                    f"Feuille '{sheet_name}' : colonne(s) manquante(s) : {', '.join(missing)}. "
                    f"Trouvées : {list(header_map.keys())}"
                ),
            )
        )
        return results, 0, 0

    for offset, row in enumerate(rows[data_start:], start=data_start + 1):
        # offset = numéro de ligne 1-indexed Excel
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def get(key: str) -> str | None:
            i = header_map.get(key)
            if i is None or i >= len(row):
                return None
            v = row[i]
            return str(v).strip() if v is not None and str(v).strip() != "" else None

        matricule_raw = get("matricule")
        if not matricule_raw:
            # Ligne sans matricule = sub-header / ligne décorative → on skip silencieusement
            continue

        matricule = matricule_raw.upper()
        if not is_valid_matricule(matricule):
            results.append(
                ImportRowResult(
                    row=offset, matricule=matricule, status="error",
                    message=f"[{sheet_name}] Format invalide (attendu XX-ESATICNNNNAA)",
                )
            )
            continue

        if matricule in seen_in_file:
            results.append(
                ImportRowResult(
                    row=offset, matricule=matricule, status="error",
                    message=f"[{sheet_name}] Doublon dans le fichier",
                )
            )
            continue
        seen_in_file.add(matricule)

        first_name = get("first_name")
        last_name = get("last_name")
        email = get("email")
        
        if not first_name or not last_name:
            results.append(
                ImportRowResult(
                    row=offset, matricule=matricule, status="error",
                    message=f"[{sheet_name}] Prénom et nom obligatoires",
                )
            )
            continue

        gender = _parse_gender(get("gender"))

        existing = db.query(Student).filter(Student.matricule == matricule).first()
        if existing:
            results.append(
                ImportRowResult(
                    row=offset, matricule=matricule, status="skipped",
                    message=f"[{sheet_name}] Déjà présent",
                )
            )
            skipped += 1
            continue

        if dry_run:
            results.append(
                ImportRowResult(
                    row=offset, matricule=matricule, status="ok",
                    message=f"[{sheet_name}] {first_name} {last_name} → {classroom.level} {classroom.name}",
                )
            )
            created += 1
            continue

        # Génération du code d'activation
        activation_code = secrets.token_hex(3).upper() # 6 caractères

        student = Student(
            matricule=matricule,
            first_name=first_name,
            last_name=last_name,
            email=email,
            activation_code=activation_code,
            password_hash=None,  # ← compte en attente d'activation
            role=UserRole.STUDENT,
            gender=gender,
            class_id=classroom.id,
            is_active=True,
        )
        db.add(student)

        # On N'ENVOIE PAS ici : l'envoi se fait après le commit, hors du cycle
        # requête/réponse. Envoyer en ligne rendait l'import O(n) appels SMTP
        # bloquants (des minutes pour une promo entière) et pouvait notifier
        # des étudiants dont l'insertion était ensuite annulée.
        if email:
            pending_emails.append(
                PendingActivationEmail(
                    to_email=email,
                    voter_name=f"{first_name} {last_name}",
                    activation_code=activation_code,
                )
            )

        results.append(
            ImportRowResult(
                row=offset, matricule=matricule, status="ok",
                message=f"[{sheet_name}] Créé : {first_name} {last_name} (Code envoyé à {email if email else 'aucun email'})",
            )
        )
        created += 1

    return results, created, skipped


def _dispatch_activation_emails(
    pending: list[PendingActivationEmail],
    background_tasks: "BackgroundTasks | None",
) -> None:
    if not pending:
        return
    if background_tasks is None:
        logger.info(
            "import: %s code(s) d'activation générés, aucun canal d'envoi fourni", len(pending)
        )
        return
    for item in pending:
        background_tasks.add_task(
            email_service.send_activation_code_email,
            to_email=item.to_email,
            voter_name=item.voter_name,
            activation_code=item.activation_code,
        )


def import_students(
    db: Session,
    *,
    file_bytes: bytes,
    dry_run: bool = False,
    auto_create_classes: bool = False,
    default_level: str | None = None,
    background_tasks: "BackgroundTasks | None" = None,
) -> ImportReport:
    """Importe tous les étudiants présents dans toutes les feuilles du xlsx.

    Si auto_create_classes=True et default_level fourni, les classes
    inexistantes sont créées automatiquement (utile pour bulk par niveau).

    Les codes d'activation ne sont envoyés qu'APRÈS le commit, et seulement si
    un `background_tasks` est fourni — sinon ils sont seulement journalisés.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return ImportReport(
            total=0, created=0, skipped=0, errors=1,
            rows=[ImportRowResult(row=0, status="error", message=f"Fichier illisible : {exc}")],
        )

    all_results: list[ImportRowResult] = []
    pending_emails: list[PendingActivationEmail] = []
    seen_in_file: set[str] = set()
    total_created = 0
    total_skipped = 0
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        total_rows += max(0, len(rows) - 1)

        results, created, skipped = _process_sheet(
            db,
            sheet_name=sheet_name,
            rows=rows,
            seen_in_file=seen_in_file,
            dry_run=dry_run,
            auto_create_classes=auto_create_classes,
            default_level=default_level,
            pending_emails=pending_emails,
        )
        all_results.extend(results)
        total_created += created
        total_skipped += skipped

    if dry_run:
        db.rollback()  # annule aussi les classes auto-créées en dry-run
        pending_emails.clear()  # rien n'a été créé, personne ne doit être notifié
    else:
        db.commit()
        _dispatch_activation_emails(pending_emails, background_tasks)

    errors = sum(1 for r in all_results if r.status == "error")
    return ImportReport(
        total=total_rows,
        created=total_created,
        skipped=total_skipped,
        errors=errors,
        rows=all_results,
    )
